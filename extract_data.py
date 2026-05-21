# -*- coding: utf-8 -*-
"""只负责读xlsx生成data.json，不做git操作"""
import os, sys, json, shutil, tempfile
from datetime import datetime
from collections import defaultdict

XLSX_SOURCE = r"C:\Users\Administrator\OneDrive\DF Docs Sharing\0.Global数据监控\1.海外分地区SOL玩家趋势（双周更新）.xlsx"
REPO_DIR    = r"C:\Users\Administrator\Desktop\df-sol-dashboard"
DATA_FILE   = os.path.join(REPO_DIR, "data.json")

TARGETS = ['Russia','United States','Japan','Germany','Taiwan, China','Thailand','Indonesia','Vietnam',
           '俄罗斯','美国','日本','德国','中国台湾','泰国','印度尼西亚','越南']
TARGETS_ZH = {'Russia':'俄罗斯','United States':'美国','Japan':'日本','Germany':'德国',
               'Taiwan, China':'中国台湾','Thailand':'泰国','Indonesia':'印度尼西亚','Vietnam':'越南'}
TARGETS_FINAL = ['俄罗斯','美国','日本','德国','中国台湾','泰国','印度尼西亚','越南']
PLATFORM_MAP     = {'手游':'手游','PC':'PC','端游':'PC','主机':'主机'}
PLATFORM_MAP_RET = {'手游注册':'手游','PC注册':'PC','端游注册':'PC','主机注册':'主机'}

def sn(x):
    if x is None: return None
    if isinstance(x, str) and x in ('-','#N/A',''): return None
    try: return round(float(x), 6)
    except: return None

def ts():
    return datetime.now().strftime('%H:%M:%S')

import openpyxl

print(f"[{ts()}] 复制xlsx到临时路径...")
tmp = tempfile.mkdtemp()
tmp_f = os.path.join(tmp, "t.xlsx")
shutil.copy2(XLSX_SOURCE, tmp_f)
print(f"[{ts()}] 复制完成，开始读取...")

src_mtime = os.path.getmtime(XLSX_SOURCE)
sol = defaultdict(lambda: defaultdict(dict))
ret = defaultdict(lambda: defaultdict(dict))

try:
    wb = openpyxl.load_workbook(tmp_f, read_only=True, data_only=True)

    print(f"[{ts()}] 读取 Sol参与情况_RD (约25万行)...")
    ws1 = wb['Sol参与情况_RD']
    count = 0
    for row in ws1.iter_rows(min_row=2, values_only=True):
        v = list(row)
        c = v[6]
        if c not in TARGETS_FINAL: continue
        p = PLATFORM_MAP.get(v[3])
        if not p or not v[2]: continue
        ds = v[2].strftime('%Y-%m-%d') if isinstance(v[2], datetime) else str(v[2])[:10]
        sol[p][c][ds] = {
            'SOL参与率': sn(v[8]), 'SOL活跃用户数': sn(v[9]),
            'SOL对局数': sn(v[10]), 'SOL人均对局数': sn(v[11]),
            'SOL人均时长（分）': sn(v[12])
        }
        count += 1
        if count % 50000 == 0:
            print(f"[{ts()}]   已处理 {count} 行...")
    print(f"[{ts()}] Sol参与情况_RD 完成，有效行: {count}")

    print(f"[{ts()}] 读取 模式留存_源数据 (约54万行)...")
    ws2 = wb['模式留存_源数据']
    count = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        v = list(row)
        c = v[5]
        if c not in TARGETS_FINAL: continue
        p = PLATFORM_MAP_RET.get(v[4])
        if not p or not v[6]: continue
        ds = v[6].strftime('%Y-%m-%d') if isinstance(v[6], datetime) else str(v[6])[:10]
        ret[p][c][ds] = {
            '模式新进量级': sn(v[7]), '模式次留': sn(v[8]),
            '模式7留': sn(v[9]), '模式30留': sn(v[10])
        }
        count += 1
        if count % 100000 == 0:
            print(f"[{ts()}]   已处理 {count} 行...")
    print(f"[{ts()}] 模式留存_源数据 完成，有效行: {count}")

    wb.close()
finally:
    shutil.rmtree(tmp, ignore_errors=True)

print(f"[{ts()}] 合并数据...")
pdata = {}
for plat in ['手游', 'PC', '主机']:
    regions = []
    for c in TARGETS_FINAL:
        s = sol[plat].get(c, {})
        r = ret[plat].get(c, {})
        dates = sorted(set(list(s.keys()) + list(r.keys())))
        recs = []
        for d in dates:
            rec = {'日期': d}
            if d in s: rec.update(s[d])
            if d in r: rec.update(r[d])
            recs.append(rec)
        regions.append({'country': c, 'data': recs})
        print(f"  {plat}-{c}: {len(recs)}期")
    pdata[plat] = regions

result = {
    'update_time': datetime.fromtimestamp(src_mtime).strftime('%Y-%m-%d'),
    'build_time':  datetime.now().strftime('%Y-%m-%d %H:%M'),
    'platforms':   pdata
}

print(f"[{ts()}] 写入 data.json...")
with open(DATA_FILE, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, separators=(',', ':'))

size_mb = os.path.getsize(DATA_FILE) / 1048576
print(f"[{ts()}] 完成！data.json 大小: {size_mb:.1f}MB")
print("下一步：运行 push_to_github.py 推送到GitHub")

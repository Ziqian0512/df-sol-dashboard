# -*- coding: utf-8 -*-
"""
sync_data.py
从xlsx读取两个sheet数据 → 生成data.json → git push到GitHub
使用方法：双击运行，或 python sync_data.py
"""
import os, sys, json, shutil, tempfile, subprocess
from datetime import datetime
from collections import defaultdict

# ============ 配置 ============
XLSX_SOURCE = r"C:\Users\Administrator\OneDrive\DF Docs Sharing\0.Global数据监控\1.海外分地区SOL玩家趋势（双周更新）.xlsx"
REPO_DIR    = r"C:\Users\Administrator\Desktop\df-sol-dashboard"
DATA_FILE   = os.path.join(REPO_DIR, "data.json")
GIT_PATH    = r"C:\Program Files\Git\cmd\git.exe"

TARGETS = ['俄罗斯','美国','日本','德国','中国台湾','泰国','印度尼西亚','越南']
PLATFORM_MAP     = {'手游':'手游','PC':'PC','端游':'PC','主机':'主机'}
PLATFORM_MAP_RET = {'手游注册':'手游','PC注册':'PC','端游注册':'PC','主机注册':'主机'}
# ==============================

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def sn(x):
    if x is None: return None
    if isinstance(x, str) and x in ('-','#N/A',''): return None
    try: return round(float(x), 6)
    except: return None

def read_xlsx():
    try:
        import openpyxl
    except ImportError:
        log("安装 openpyxl...")
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
        import openpyxl

    log(f"复制xlsx到临时路径（避免OneDrive锁文件）...")
    tmp = tempfile.mkdtemp()
    tmp_f = os.path.join(tmp, "t.xlsx")
    shutil.copy2(XLSX_SOURCE, tmp_f)

    try:
        wb = openpyxl.load_workbook(tmp_f, read_only=True, data_only=True)

        # --- Sol参与情况_RD ---
        log("读取 Sol参与情况_RD...")
        ws1 = wb['Sol参与情况_RD']
        sol = defaultdict(lambda: defaultdict(dict))
        count = 0
        for row in ws1.iter_rows(min_row=2, values_only=True):
            v = list(row)
            c = v[6]
            if c not in TARGETS: continue
            p = PLATFORM_MAP.get(v[3])
            if not p or not v[2]: continue
            ds = v[2].strftime('%Y-%m-%d') if isinstance(v[2], datetime) else str(v[2])[:10]
            sol[p][c][ds] = {
                'SOL参与率':       sn(v[8]),
                'SOL活跃用户数':   sn(v[9]),
                'SOL对局数':       sn(v[10]),
                'SOL人均对局数':   sn(v[11]),
                'SOL人均时长（分）': sn(v[12])
            }
            count += 1
        log(f"  有效行: {count}")

        # --- 模式留存_源数据 ---
        log("读取 模式留存_源数据...")
        ws2 = wb['模式留存_源数据']
        ret = defaultdict(lambda: defaultdict(dict))
        count = 0
        for row in ws2.iter_rows(min_row=2, values_only=True):
            v = list(row)
            c = v[5]
            if c not in TARGETS: continue
            p = PLATFORM_MAP_RET.get(v[4])
            if not p or not v[6]: continue
            ds = v[6].strftime('%Y-%m-%d') if isinstance(v[6], datetime) else str(v[6])[:10]
            ret[p][c][ds] = {
                '模式新进量级': sn(v[7]),
                '模式次留':     sn(v[8]),
                '模式7留':      sn(v[9]),
                '模式30留':     sn(v[10])
            }
            count += 1
        log(f"  有效行: {count}")

        wb.close()
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    # --- 合并 ---
    log("合并数据...")
    src_mtime = os.path.getmtime(XLSX_SOURCE)
    pdata = {}
    for plat in ['手游', 'PC', '主机']:
        regions = []
        for c in TARGETS:
            s = sol[plat].get(c, {})
            r = ret[plat].get(c, {})
            dates = sorted(set(list(s.keys()) + list(r.keys())))
            recs = []
            for d in dates:
                rec = {'日期': d}
                if d in s: rec.update(s[d])
                if d in r: rec.update(r[d])
                recs.append(rec)
            regions.append({'country': c, 'data': recs})
            if recs: log(f"  {plat}-{c}: {len(recs)}期")
        pdata[plat] = regions

    result = {
        'update_time': datetime.fromtimestamp(src_mtime).strftime('%Y-%m-%d'),
        'build_time':  datetime.now().strftime('%Y-%m-%d %H:%M'),
        'platforms':   pdata
    }
    return result

def git_push(data):
    log("写入 data.json...")
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    size_mb = os.path.getsize(DATA_FILE) / 1048576
    log(f"data.json 大小: {size_mb:.1f}MB")

    if size_mb > 90:
        log("警告：文件超过90MB，GitHub单文件限制100MB，请注意！")

    def git(args):
        result = subprocess.run(
            [GIT_PATH] + args,
            cwd=REPO_DIR, capture_output=True, text=True, encoding='utf-8'
        )
        if result.stdout.strip(): log(f"  git: {result.stdout.strip()}")
        if result.stderr.strip(): log(f"  git: {result.stderr.strip()}")
        return result.returncode

    log("拉取远端最新...")
    git(['branch', '-M', 'main'])
    git(['pull', 'origin', 'main', '--rebase', '--allow-unrelated-histories'])

    log("添加文件...")
    git(['add', 'data.json'])

    commit_msg = f"数据更新 {data['update_time']} (build: {data['build_time']})"
    log(f"提交: {commit_msg}")
    git(['commit', '-m', commit_msg])

    log("推送到GitHub...")
    ret = git(['push', 'origin', 'main'])
    if ret == 0:
        log("[OK] 推送成功！")
        log(f"数据已更新：https://Ziqian0512.github.io/df-sol-dashboard/")
    else:
        log("[FAIL] 推送失败，请检查网络或token是否有效")

if __name__ == '__main__':
    log("=" * 50)
    log("SOL数据同步脚本")
    log("=" * 50)
    try:
        data = read_xlsx()
        git_push(data)
    except Exception as e:
        log(f"[ERROR] {e}")
        import traceback; traceback.print_exc()
    log("=" * 50)
    input("按回车键退出...")
